from datetime import datetime
from typing import List

from openpyxl import load_workbook
from sqlmodel import Session, select

import settings
from db import get_engine
from models.base import DoLDataSource
from models.dol_disclosure_job_order import DolDisclosureJobOrder

valid_col_names = (
    "case_number",
    "case_status",
    "received_date",
    "decision_date",
    "type_of_employer_application",
    "h2a_labor_contractor",
    "nature_of_temporary_need",
    "emergency_filing",
    "employer_name",
    "trade_name_dba",
    "employer_address_1",
    "employer_address_2",
    "employer_city",
    "employer_state",
    "employer_postal_code",
    "employer_country",
    "employer_province",
    "employer_phone",
    "employer_phone_ext",
    "naics_code",
    "employer_poc_last_name",
    "employer_poc_first_name",
    "employer_poc_middle_name",
    "employer_poc_job_title",
    "employer_poc_address1",
    "employer_poc_address2",
    "employer_poc_city",
    "employer_poc_state",
    "employer_poc_postal_code",
    "employer_poc_country",
    "employer_poc_province",
    "employer_poc_phone",
    "employer_poc_phone_ext",
    "employer_poc_email",
    "type_of_representation",
    "attorney_agent_last_name",
    "attorney_agent_first_name",
    "attorney_agent_middle_name",
    "attorney_agent_address_1",
    "attorney_agent_address_2",
    "attorney_agent_city",
    "attorney_agent_state",
    "attorney_agent_postal_code",
    "attorney_agent_country",
    "attorney_agent_province",
    "attorney_agent_phone",
    "attorney_agent_phone_ext",
    "attorney_agent_email_address",
    "lawfirm_name_business_name",
    "state_of_highest_court",
    "name_of_highest_state_court",
    "soc_code",
    "soc_title",
    "seven_ninety_a_addendum_b_attached",
    "work_contracts_attached",
    "employer_mspa_attached",
    "surety_bond_attached",
    "housing_transportation",
    "appendix_a_attached",
    "jnt_employer_append_a_attached",
    "preparer_last_name",
    "preparer_first_name",
    "preparer_middle_initial",
    "preparer_business_name",
    "preparer_email",
    "job_order_number",
    "job_title",
    "total_workers_needed",
    "total_workers_h2a_requested",
    "total_workers_h2a_certified",
    "requested_begin_date",
    "requested_end_date",
    "employment_begin_date",
    "employment_end_date",
    "on_call_requirement",
    "anticipated_number_of_hours",
    "sunday_hours",
    "monday_hours",
    "tuesday_hours",
    "wednesday_hours",
    "thursday_hours",
    "friday_hours",
    "saturday_hours",
    "hourly_schedule_begin",
    "hourly_schedule_end",
    "wage_offer",
    "per",
    "piece_rate_offer",
    "piece_rate_unit",
    "seven_ninety_a_addendum_a_attached",
    "frequency_of_pay",
    "other_frequency_of_pay",
    "deductions_from_pay",
    "education_level",
    "work_experience_months",
    "training_months",
    "certification_requirements",
    "driver_requirements",
    "criminal_background_check",
    "drug_screen",
    "lifting_requirements",
    "lifting_amount",
    "exposure_to_temperatures",
    "extensive_pushing_pulling",
    "extensive_sitting_walking",
    "frequent_stooping_bending_over",
    "repetitive_movements",
    "supervise_other_emp",
    "supervise_how_many",
    "special_requirements",
    "worksite_address",
    "worksite_city",
    "worksite_state",
    "worksite_postal_code",
    "worksite_county",
    "addendum_b_worksite_attached",
    "total_worksites_records",
    "housing_address_location",
    "housing_city",
    "housing_state",
    "housing_postal_code",
    "housing_county",
    "type_of_housing",
    "total_units",
    "total_occupancy",
    "housing_compliance_local",
    "housing_compliance_state",
    "housing_compliance_federal",
    "addendum_b_housing_attached",
    "total_housing_records",
    "meals_provided",
    "meals_charge",
    "meal_reimbursement_minimum",
    "meal_reimbursement_maximum",
    "phone_to_apply",
    "email_to_apply",
    "website_to_apply",
    "addendum_c_attached",
    "total_addendum_a_records",
)

alternate_col_names = {
    "790a_addendum_b_attached": "seven_ninety_a_addendum_b_attached",
    "790a_addendum_a_attached": "seven_ninety_a_addendum_a_attached",
    "employer_address1": "employer_address_1",
    "employer_address2": "employer_address_2",
    "case_no": "case_number",
    "case_received_date": "received_date",
    # TODO: Populate this with reference to the existing DoL filings over the years
    "agent_attorney_city": "attorney_agent_city",
    "agent_attorney_state": "attorney_agent_state",
    "nbr_workers_requested": "total_workers_h2a_requested",
    "nbr_workers_certified": "total_workers_h2a_certified",
    "basic_number_of_hours": "anticipated_number_of_hours",
    "basic_rate_of_pay": "wage_offer",
    "basic_unit_of_pay": "per",
}


def row_to_dict(header_row: List[str], row: List):
    output_dict = {}
    for i, k in enumerate(header_row):
        if k:
            output_dict[k] = row[i]
    return output_dict


def import_disclosure(filename: str):
    """
    Import a DoL disclosure file spreadsheet
    :param filename: Filename to import
    :return:
    """

    # Check if this has already been imported and exit if it has been.
    session = Session(get_engine())
    exist_rows = session.exec(
        select(DolDisclosureJobOrder).where(DolDisclosureJobOrder.file_name == filename)
    ).first()
    if exist_rows:
        print(f"Filename {filename} has already been imported! Quitting.")
        return

    wb = load_workbook(
        filename=filename, read_only=True, keep_links=False, data_only=True
    )
    worksheet = wb.active

    # Process header row and filter out for only valid column names.
    header_row = worksheet[1]
    col_names = [str(c.value).lower() for c in header_row]
    for i, name in enumerate(col_names):
        if name not in valid_col_names:
            if name not in alternate_col_names:
                print(f"'{name}': '',")
            col_names[i] = alternate_col_names.get(name)

    count = 0
    pub_date = datetime.utcnow

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = row_to_dict(col_names, row)

        session.add(
            DolDisclosureJobOrder(
                source=DoLDataSource.dol_disclosure,
                pub_date=pub_date,
                file_name=filename,
                file_row=count + 1,
                first_seen=values["received_date"],
                last_seen=values["received_date"],
                **values,
            ).clean()
        )
        count += 1

        if count % settings.ROWS_BEFORE_COMMIT == 0:
            session.commit()

    session.commit()
    print(f"{count} listings imported from file {filename}")


if __name__ == "__main__":
    import_disclosure("../files/H-2A_Disclosure_Data_FY2021.xlsx")
