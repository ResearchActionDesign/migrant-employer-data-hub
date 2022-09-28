import io
from sys import stderr
from typing import List, Union

import boto3
from openpyxl import load_workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app import settings
from app.db import get_engine
from app.models.base import DoLDataSource
from app.models.dol_disclosure_job_order import DolDisclosureJobOrder
from app.models.imported_dataset import ImportedDataset, ImportStatus

valid_col_names = (
    "case_number",
    "case_status",
    "received_date",
    "decision_date",
    "type_of_employer_application",
    "h2a_labor_contractor",
    "nature_of_temporary_need",
    "emergency_filing",
    "employer_name",
    "trade_name_dba",
    "employer_address_1",
    "employer_address_2",
    "employer_city",
    "employer_state",
    "employer_postal_code",
    "employer_country",
    "employer_province",
    "employer_phone",
    "employer_phone_ext",
    "naics_code",
    "employer_poc_last_name",
    "employer_poc_first_name",
    "employer_poc_middle_name",
    "employer_poc_job_title",
    "employer_poc_address1",
    "employer_poc_address2",
    "employer_poc_city",
    "employer_poc_state",
    "employer_poc_postal_code",
    "employer_poc_country",
    "employer_poc_province",
    "employer_poc_phone",
    "employer_poc_phone_ext",
    "employer_poc_email",
    "type_of_representation",
    "attorney_agent_last_name",
    "attorney_agent_first_name",
    "attorney_agent_middle_name",
    "attorney_agent_address_1",
    "attorney_agent_address_2",
    "attorney_agent_city",
    "attorney_agent_state",
    "attorney_agent_postal_code",
    "attorney_agent_country",
    "attorney_agent_province",
    "attorney_agent_phone",
    "attorney_agent_phone_ext",
    "attorney_agent_email_address",
    "lawfirm_name_business_name",
    "state_of_highest_court",
    "name_of_highest_state_court",
    "soc_code",
    "soc_title",
    "seven_ninety_a_addendum_b_attached",
    "work_contracts_attached",
    "employer_mspa_attached",
    "surety_bond_attached",
    "housing_transportation",
    "appendix_a_attached",
    "jnt_employer_append_a_attached",
    "preparer_last_name",
    "preparer_first_name",
    "preparer_middle_initial",
    "preparer_business_name",
    "preparer_email",
    "job_order_number",
    "job_title",
    "total_workers_needed",
    "total_workers_h2a_requested",
    "total_workers_h2a_certified",
    "requested_begin_date",
    "requested_end_date",
    "employment_begin_date",
    "employment_end_date",
    "on_call_requirement",
    "anticipated_number_of_hours",
    "sunday_hours",
    "monday_hours",
    "tuesday_hours",
    "wednesday_hours",
    "thursday_hours",
    "friday_hours",
    "saturday_hours",
    "hourly_schedule_begin",
    "hourly_schedule_end",
    "wage_offer",
    "per",
    "piece_rate_offer",
    "piece_rate_unit",
    "seven_ninety_a_addendum_a_attached",
    "frequency_of_pay",
    "other_frequency_of_pay",
    "deductions_from_pay",
    "education_level",
    "work_experience_months",
    "training_months",
    "certification_requirements",
    "driver_requirements",
    "criminal_background_check",
    "drug_screen",
    "lifting_requirements",
    "lifting_amount",
    "exposure_to_temperatures",
    "extensive_pushing_pulling",
    "extensive_sitting_walking",
    "frequent_stooping_bending_over",
    "repetitive_movements",
    "supervise_other_emp",
    "supervise_how_many",
    "special_requirements",
    "worksite_address",
    "worksite_city",
    "worksite_state",
    "worksite_postal_code",
    "worksite_county",
    "addendum_b_worksite_attached",
    "total_worksites_records",
    "housing_address_location",
    "housing_city",
    "housing_state",
    "housing_postal_code",
    "housing_county",
    "type_of_housing",
    "total_units",
    "total_occupancy",
    "housing_compliance_local",
    "housing_compliance_state",
    "housing_compliance_federal",
    "addendum_b_housing_attached",
    "total_housing_records",
    "meals_provided",
    "meals_charge",
    "meal_reimbursement_minimum",
    "meal_reimbursement_maximum",
    "phone_to_apply",
    "email_to_apply",
    "website_to_apply",
    "addendum_c_attached",
    "total_addendum_a_records",
)

alternate_col_names = {
    "790a_addendum_b_attached": "seven_ninety_a_addendum_b_attached",
    "790a_addendum_a_attached": "seven_ninety_a_addendum_a_attached",
    "employer_address1": "employer_address_1",
    "employer_address2": "employer_address_2",
    "case_no": "case_number",
    "case_received_date": "received_date",
    # TODO: Populate this with reference to the existing DoL filings over the years
    "agent_attorney_city": "attorney_agent_city",
    "agent_attorney_state": "attorney_agent_state",
    "nbr_workers_requested": "total_workers_h2a_requested",
    "nbr_workers_certified": "total_workers_h2a_certified",
    "basic_number_of_hours": "anticipated_number_of_hours",
    "basic_rate_of_pay": "wage_offer",
    "basic_unit_of_pay": "per",
}

s3_client = boto3.client("s3")


def row_to_dict(header_row: List[str], row: List):
    output_dict = {}
    for i, k in enumerate(header_row):
        if k:
            output_dict[k] = row[i]
    return output_dict


def import_disclosure(
    filename: Union[str, None] = None,
    bucket_name: Union[str, None] = None,
    object_name: Union[str, None] = None,
) -> bool:
    """
    Import a DoL disclosure file spreadsheet
    :param filename: Filename to import
    :return:
    """
    if not filename and (not bucket_name or not object_name):
        stderr.write("No valid parameters specified.")
        return False

    file_id = filename or object_name

    if file_id is None:
        stderr.write("No valid parameters specified.")
        return False

    # Check if this has already been imported and exit if it has been.
    session = Session(get_engine())

    visa_class = None
    if "h-2a" in file_id.lower() and "h-2b" not in file_id.lower():
        visa_class = "H-2A"
    elif "h-2b" in file_id.lower() and "h-2a" not in file_id.lower():
        visa_class = "H-2B"
    # TODO: other visa types.

    if not filename:
        f = io.BytesIO()
        s3_client.download_fileobj(Bucket=bucket_name, Key=object_name, Fileobj=f)
        wb = load_workbook(filename=f, read_only=True, keep_links=False, data_only=True)

    else:
        wb = load_workbook(
            filename=filename, read_only=True, keep_links=False, data_only=True
        )

    worksheet = wb.active

    import_count = session.exec(
        select(func.max(DolDisclosureJobOrder.file_row)).where(
            DolDisclosureJobOrder.file_name == file_id
        )
    ).first()

    if import_count and import_count + 1 >= worksheet.max_row:
        print(f"File {file_id} has already been imported! Quitting.")
        return True

    if import_count:
        print(
            f"File {file_id} has already been started, continuing partial import with row {import_count + 1}"
        )

    else:
        import_count = 0

    print(f"Importing {file_id}")

    # Process header row and filter out for only valid column names.
    header_row = worksheet[1]
    col_names = [str(c.value).lower() for c in header_row]
    for i, name in enumerate(col_names):
        if name not in valid_col_names:
            if name not in alternate_col_names:
                print("Missing column names:")
                print(f"'{name}': '',")
            col_names[i] = alternate_col_names.get(name)

    count = import_count

    for row in worksheet.iter_rows(
        min_row=(import_count + 2 if import_count else 2), values_only=True
    ):
        values = row_to_dict(col_names, row)

        session.add(
            DolDisclosureJobOrder(
                source=DoLDataSource.dol_disclosure,
                file_name=file_id,
                file_row=count + 1,
                first_seen=values["received_date"],
                last_seen=values["received_date"],
                visa_class=visa_class,
                **values,
            ).clean()
        )
        count += 1

        if count % settings.ROWS_BEFORE_COMMIT == 0:
            print(f"{count} listings imported from file {file_id}")
            session.commit()

    session.commit()
    session.close()
    print(f"{count} listings imported from file {file_id}")
    return True


def process_imports() -> bool:
    """
    Query the DB to see if there are any imports which need to be processed; and call import_disclosure if so.
    :return:
    """
    session = Session(get_engine())
    import_to_do = session.exec(
        select(ImportedDataset).where(
            ImportedDataset.import_status == ImportStatus.needs_importing
        )
    ).first()
    session.close()

    finished = False
    if import_to_do.bucket_name:
        finished = import_disclosure(
            bucket_name=import_to_do.bucket_name, object_name=import_to_do.object_name
        )
    else:
        finished = import_disclosure(filename=import_to_do.object_name)

    if finished:
        session = Session(get_engine())
        import_to_do.import_status = ImportStatus.finished
        session.add(import_to_do)
        session.commit()
        session.close()

    return True


def add_new_import(
    filename: Union[str, None] = None,
    bucket_name: Union[str, None] = None,
    object_name: Union[str, None] = None,
):
    """
    Adds a new import that needs to be done.
    :param filename:
    :param bucket_name:
    :param object_name:
    :return:
    """
    session = Session(get_engine())
    import_to_do = ImportedDataset(
        bucket_name=bucket_name, object_name=(filename or object_name)
    )
    session.add(import_to_do)
    session.commit()
    session.close()


if __name__ == "__main__":
    import_disclosure("../files/H-2A_Disclosure_Data_FY2021.xlsx")
